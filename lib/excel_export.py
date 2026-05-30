import io
import csv
from datetime import datetime
from collections import defaultdict
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


CATEGORY_MAP = {
    "office_supplies": "辦公用品 Office Supplies",
    "meals_entertainment": "膳食及應酬 Meals & Entertainment",
    "transportation": "交通費 Transportation",
    "utilities": "水電費 Utilities",
    "rent_rates": "租金及差餉 Rent & Rates",
    "professional_fees": "專業費用 Professional Fees",
    "insurance": "保險 Insurance",
    "repairs_maintenance": "維修保養 Repairs & Maintenance",
    "travel": "差旅費 Travel",
    "marketing": "市場推廣 Marketing",
    "depreciation": "折舊 Depreciation",
    "miscellaneous": "雜項 Miscellaneous",
}

RECEIPT_TYPE_LABELS = {
    "office_supplies": "辦公用品 Office Supplies",
    "meals_entertainment": "膳食及應酬 Meals & Entertainment",
    "transportation": "交通費 Transportation",
    "utilities": "水電費 Utilities",
    "rent_rates": "租金及差餉 Rent & Rates",
    "professional_fees": "專業費用 Professional Fees",
    "insurance": "保險 Insurance",
    "repairs_maintenance": "維修保養 Repairs & Maintenance",
    "travel": "差旅費 Travel",
    "marketing": "市場推廣 Marketing",
    "depreciation": "折舊 Depreciation",
    "miscellaneous": "雜項 Miscellaneous",
}


def _map_category(receipt_type: str) -> str:
    return CATEGORY_MAP.get(receipt_type, CATEGORY_MAP["miscellaneous"])


def _format_type(receipt_type: str) -> str:
    return RECEIPT_TYPE_LABELS.get(receipt_type, receipt_type)


def _fmt_hkd(val) -> str:
    try:
        return f"{float(val):,.2f}"
    except (TypeError, ValueError):
        return "0.00"


def _build_receipt_rows(receipts: list) -> list:
    rows = []
    for idx, r in enumerate(receipts, start=1):
        total = float(r.get("total_amount", 0) or 0)
        rows.append({
            "No.": idx,
            "Receipt Date 日期": r.get("receipt_date", ""),
            "Merchant Name 商戶名稱": r.get("merchant_name", ""),
            "Description 描述": r.get("notes", ""),
            "Receipt Type 收據類型": _format_type(r.get("receipt_type", "miscellaneous")),
            "Expense Category 費用類別": _map_category(r.get("receipt_type", "miscellaneous")),
            "Total Amount HKD 總額": total,
            "Payment Method 付款方式": r.get("payment_method", "N/A") or "N/A",
            "Status 狀態": r.get("status", ""),
            "Approver 審批人": r.get("approved_by", ""),
            "Remarks 備註": r.get("rejection_reason", ""),
        })
    return rows


def _style_sheet(ws, df):
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, len(df.columns) + 1):
        max_len = max(
            len(str(df.columns[col_idx - 1])),
            *(len(str(df.iloc[row, col_idx - 1])) for row in range(min(len(df), 100))),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


def generate_excel_report(receipts: list) -> bytes:
    if not receipts:
        return None

    try:
        rows = _build_receipt_rows(receipts)
        df_receipts = pd.DataFrame(rows)

        hkd_cols = [
            "Total Amount HKD 總額",
        ]
        for col in hkd_cols:
            if col in df_receipts.columns:
                df_receipts[col] = df_receipts[col].apply(lambda v: _fmt_hkd(v))

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_receipts.to_excel(writer, sheet_name="Receipts 收據", index=False)
            ws = writer.sheets["Receipts 收據"]
            _style_sheet(ws, df_receipts)

            total_amt = sum(float(r.get("total_amount", 0) or 0) for r in receipts)

            dates = [r.get("receipt_date", "") for r in receipts if r.get("receipt_date")]
            period = ""
            if dates:
                period = f"{min(dates)} to {max(dates)}"

            summary_data = {
                "Metric": [
                    "Total Receipts 收據總數",
                    "Total Amount HKD 總額",
                    "Period 期間",
                ],
                "Value": [
                    str(len(receipts)),
                    _fmt_hkd(total_amt),
                    period,
                ],
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name="Summary 摘要", index=False)
            ws_summary = writer.sheets["Summary 摘要"]
            _style_sheet(ws_summary, df_summary)

            cat_totals = defaultdict(float)
            cat_counts = defaultdict(int)
            for r in receipts:
                cat = _map_category(r.get("receipt_type", "miscellaneous"))
                cat_totals[cat] += float(r.get("total_amount", 0) or 0)
                cat_counts[cat] += 1

            cat_rows = []
            for cat in sorted(cat_totals.keys()):
                pct = (cat_totals[cat] / total_amt * 100) if total_amt else 0
                cat_rows.append({
                    "Expense Category 費用類別": cat,
                    "Count 數量": cat_counts[cat],
                    "Total Amount HKD 總額": _fmt_hkd(cat_totals[cat]),
                    "Percentage of Total 佔比": f"{pct:.1f}%",
                })
            df_cat = pd.DataFrame(cat_rows)
            df_cat.to_excel(writer, sheet_name="Category Summary 類別摘要", index=False)
            ws_cat = writer.sheets["Category Summary 類別摘要"]
            _style_sheet(ws_cat, df_cat)

            month_cat = defaultdict(lambda: defaultdict(lambda: {"count": 0, "amount": 0.0}))
            for r in receipts:
                month = (r.get("receipt_date", "") or "")[:7]
                if not month:
                    month = "Unknown"
                cat = _map_category(r.get("receipt_type", "miscellaneous"))
                month_cat[month][cat]["count"] += 1
                month_cat[month][cat]["amount"] += float(r.get("total_amount", 0) or 0)

            month_rows = []
            for month in sorted(month_cat.keys()):
                for cat in sorted(month_cat[month].keys()):
                    month_rows.append({
                        "Month 月份": month,
                        "Category 類別": cat,
                        "Count 數量": month_cat[month][cat]["count"],
                        "Total Amount HKD 總額": _fmt_hkd(month_cat[month][cat]["amount"]),
                    })
            df_month = pd.DataFrame(month_rows)
            df_month.to_excel(writer, sheet_name="Monthly Breakdown 月度分析", index=False)
            ws_month = writer.sheets["Monthly Breakdown 月度分析"]
            _style_sheet(ws_month, df_month)

        return output.getvalue()
    except Exception:
        return None


def generate_csv_report(receipts: list) -> str:
    if not receipts:
        return None

    try:
        rows = _build_receipt_rows(receipts)
        if not rows:
            return None

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=rows[0].keys())
        writer.writeheader()
        for row in rows:
            formatted = {}
            for k, v in row.items():
                if k in ("Total Amount HKD 總額", "Tax Amount HKD 稅款", "Net Amount HKD 淨額"):
                    formatted[k] = _fmt_hkd(v)
                else:
                    formatted[k] = v
            writer.writerow(formatted)

        return output.getvalue()
    except Exception:
        return None
